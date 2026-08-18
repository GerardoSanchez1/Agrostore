import os
import sys
import django
import shutil
from pathlib import Path
import openpyxl

# Set up Django environment
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'mystore.settings')
django.setup()

from store.models import Brand, Category, Item, Picture
from django.core.files import File
from django.conf import settings

BASE_DIR = Path(__file__).resolve().parent
INFO_DIR = BASE_DIR / 'Information'
EXCEL_FILE = INFO_DIR / 'Information.xlsx'
MEDIA_SRC_DIR = INFO_DIR / 'media'
TEXT_SRC_DIR = INFO_DIR / 'text'

def main():
    if not EXCEL_FILE.exists():
        print(f"Error: Could not find {EXCEL_FILE}")
        sys.exit(1)

    print("Loading workbook...")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    
    # Import Brands
    print("Importing Brands...")
    if 'Brand' in wb.sheetnames:
        for row in wb['Brand'].iter_rows(min_row=2, values_only=True):
            if not row or not row[0]: continue
            b_id, name, logo_file = row[0], row[1], row[2]
            logo_file = f"{logo_file}.jpg" if logo_file and not str(logo_file).endswith('.jpg') else logo_file
            
            brand, created = Brand.objects.update_or_create(
                id=b_id, 
                defaults={'name': str(name)}
            )
            
            if logo_file:
                src_path = MEDIA_SRC_DIR / logo_file
                if src_path.exists():
                    with open(src_path, 'rb') as f:
                        brand.logo.save(logo_file, File(f), save=True)
            print(f"  {'Created' if created else 'Updated'} Brand: {name}")

    # Import Categories
    print("Importing Categories...")
    if 'Category' in wb.sheetnames:
        for row in wb['Category'].iter_rows(min_row=2, values_only=True):
            if not row or not row[0]: continue
            c_id, name = row[0], row[1]
            cat, created = Category.objects.update_or_create(
                id=c_id, 
                defaults={'name': str(name)}
            )
            print(f"  {'Created' if created else 'Updated'} Category: {name}")
                
    # Import Items
    print("Importing Items...")
    if 'Item' in wb.sheetnames:
        for row in wb['Item'].iter_rows(min_row=2, values_only=True):
            if not row or not row[0]: continue
            
            i_id, name, b_id, c_id, price, desc_file, details_file = row[0:7]
            
            description = ""
            details = ""
            
            if desc_file:
                df_path = TEXT_SRC_DIR / f"{desc_file}.txt"
                if df_path.exists():
                    with open(df_path, 'r', encoding='utf-8') as f:
                        description = f.read()
                        
            if details_file:
                dt_path = TEXT_SRC_DIR / f"{details_file}.txt"
                if dt_path.exists():
                    with open(dt_path, 'r', encoding='utf-8') as f:
                        details = f.read()

            try:
                brand = Brand.objects.get(id=b_id)
                category = Category.objects.get(id=c_id)
                
                item, created = Item.objects.update_or_create(
                    id=i_id,
                    defaults={
                        'name': str(name),
                        'brand': brand,
                        'category': category,
                        'price': price if price is not None else 0.0,
                        'description': description,
                        'details': details
                    }
                )
                print(f"  {'Created' if created else 'Updated'} Item: {name}")
            except Exception as e:
                print(f"  Error with item {name}: {e}")

    # Import Pictures
    print("Importing Pictures...")
    if 'Picture' in wb.sheetnames:
        for row in wb['Picture'].iter_rows(min_row=2, values_only=True):
            if not row or not row[0]: continue
            
            p_id, photo_file, i_id = row[0], row[1], row[2]
            photo_file = f"{photo_file}.jpg" if photo_file and not str(photo_file).endswith('.jpg') else photo_file
            
            if i_id and photo_file:
                try:
                    item = Item.objects.get(id=i_id)
                    src_path = MEDIA_SRC_DIR / photo_file
                    
                    if src_path.exists():
                        pic, created = Picture.objects.get_or_create(id=p_id, item=item)
                        with open(src_path, 'rb') as f:
                            pic.photo.save(photo_file, File(f), save=True)
                        print(f"  Added picture {photo_file} to {item.name}")
                    else:
                        print(f"  Image not found: {src_path}")
                except Item.DoesNotExist:
                    print(f"  Item not found for picture: {p_id}")

    print("Import complete!")

if __name__ == '__main__':
    main()
